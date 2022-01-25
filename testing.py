import openpyxl as oxl

data = {'thing1': 'thing 1 property', 'thing2': 'thing2 property'}

workbook = oxl.Workbook()
sheet = workbook.active

def main(pt):
    workbook = oxl.Workbook()
    sheet = workbook.active
    if sheet.max_row == 1:
        cols = 1
        for header in list(pt.keys()):
            sheet.cell(1, cols, header)
            cols += 1

    for key, value in pt.items():
        print(value)
        sheet.cell(sheet.max_row + 1, 1, value)
    workbook.save('write2cell.xlsx')
main(data)