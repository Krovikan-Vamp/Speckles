import openpyxl as oxl

wb = oxl.load_workbook('Requests.xlsx')
ws = wb.active

content = {
    "name": 'I should be in the first col', 
    "what": 'row 6 coming up',
    "who":  'thin12213g3'
}

rTW = ws.max_row + 1
col = 1

for key, value in content.items():
    ws.cell(rTW, col, value)
    col += 1

wb.save('Requests.xlsx')